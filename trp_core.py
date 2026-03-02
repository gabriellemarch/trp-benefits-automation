# trp_core.py
from __future__ import annotations

import calendar
import hashlib
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import pandas as pd

import os
import calendar
from datetime import date
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------
# Config
# ----------------------------

FISCAL_QUARTERS = {
    "Q1": [4, 5, 6],      # Apr–Jun
    "Q2": [7, 8, 9],      # Jul–Sep
    "Q3": [10, 11, 12],   # Oct–Dec
    "Q4": [1, 2, 3],      # Jan–Mar
}

PROGRAM_RBW = "RBW"
PROGRAM_CARPOOL = "CARPOOL"
PROGRAM_RAD = "RAD"
PROGRAM_AFV = "AFV"

EXPECTED_OUTPUT_FILES = [
    "cleaned_master.csv",
    "lunch_report.csv",
    "winner_report.csv",
    "run_log.json",
]

# ----------------------------
# Helpers: reading files
# ----------------------------

def read_table(path: str) -> pd.DataFrame:
    """
    Reads CSV or Excel based on extension.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = p.suffix.lower()
    if ext in [".csv"]:
        return pd.read_csv(path)
    if ext in [".xlsx", ".xls"]:
        return pd.read_excel(path)
    # fallback: try csv
    return pd.read_csv(path)

# ----------------------------
# Helpers: cleaning
# ----------------------------

def _safe_str(x) -> str:
    if pd.isna(x):
        return ""
    return str(x)

def clean_email(email: str) -> str:
    email = _safe_str(email).strip().lower()
    email = re.sub(r"\s+", "", email)
    return email

def clean_badge(badge: str) -> str:
    badge = _safe_str(badge).strip().upper()
    badge = re.sub(r"[^A-Z0-9]", "", badge)
    return badge

def smart_title(name: str) -> str:
    """
    Title-case but keep apostrophes/hyphens reasonably intact.
    """
    name = _safe_str(name).strip()
    name = re.sub(r"\s+", " ", name)

    def _fix_token(tok: str) -> str:
        parts = re.split(r"([\'-])", tok)
        parts = [p.capitalize() if p not in ["'", "-"] else p for p in parts]
        return "".join(parts)

    tokens = name.split(" ")
    tokens = [_fix_token(t) for t in tokens if t]
    return " ".join(tokens)

def parse_datetime(series: pd.Series) -> pd.Series:
    """
    Robust datetime parsing. Coerces invalid to NaT.
    """
    return pd.to_datetime(series, errors="coerce")

# ----------------------------
# Loaders: map each file to standard columns
# ----------------------------

def load_rbw(path: str) -> pd.DataFrame:
    df = read_table(path)
    out = pd.DataFrame({
        "name_raw": df.get("Name"),
        "badge_raw": df.get("Badge ID Number"),
        "email_raw": df.get("Email"),
        "created_raw": df.get("Created"),
    })
    out["program"] = PROGRAM_RBW
    out["created_date"] = parse_datetime(out["created_raw"])
    return out

def load_carpool(path: str) -> pd.DataFrame:
    df = read_table(path)
    out = pd.DataFrame({
        "name_raw": df.get("Name"),
        "badge_raw": df.get("Badge ID Number"),
        "email_raw": df.get("Email"),
        "created_raw": df.get("Created"),
    })
    out["program"] = PROGRAM_CARPOOL
    out["created_date"] = parse_datetime(out["created_raw"])
    return out

def load_rad(path: str) -> pd.DataFrame:
    df = read_table(path)
    out = pd.DataFrame({
        "name_raw": df.get("Name"),
        "badge_raw": df.get("Badge ID Number"),
        "email_raw": df.get("Microchip Email"),
        "created_raw": df.get("Refuel Date and Time"),
    })
    out["program"] = PROGRAM_RAD
    out["created_date"] = parse_datetime(out["created_raw"])
    return out

def load_afv(path: str) -> pd.DataFrame:
    df = read_table(path)
    out = pd.DataFrame({
        "name_raw": df.get("Name"),
        "badge_raw": df.get("Badge #"),
        "email_raw": df.get("Email"),
        "created_raw": pd.Series([None] * len(df)),
    })
    out["program"] = PROGRAM_AFV
    out["created_date"] = pd.NaT
    return out

def fiscal_quarter_label(dt: Optional[pd.Timestamp]) -> Optional[str]:
    if pd.isna(dt):
        return None
    m = int(dt.month)
    y = int(dt.year)

    if m in FISCAL_QUARTERS["Q1"]:
        return f"Q1-{y}"
    if m in FISCAL_QUARTERS["Q2"]:
        return f"Q2-{y}"
    if m in FISCAL_QUARTERS["Q3"]:
        return f"Q3-{y}"
    if m in FISCAL_QUARTERS["Q4"]:
        return f"Q4-{y}"
    return None

def standardize(master: pd.DataFrame) -> pd.DataFrame:
    master = master.copy()
    master["name"] = master["name_raw"].apply(smart_title)
    master["badge_id"] = master["badge_raw"].apply(clean_badge)
    master["email"] = master["email_raw"].apply(clean_email)

    master["has_badge"] = master["badge_id"].astype(str).str.len() > 0
    master["has_email"] = master["email"].astype(str).str.contains("@", na=False)

    master["year"] = master["created_date"].dt.year
    master["month"] = master["created_date"].dt.month
    master["ym"] = master["created_date"].dt.to_period("M").astype(str)

    master["fiscal_q"] = master.apply(lambda r: fiscal_quarter_label(r["created_date"]), axis=1)

    cols = [
        "program",
        "name", "badge_id", "email",
        "created_date", "year", "month", "ym", "fiscal_q",
        "has_badge", "has_email",
        "name_raw", "badge_raw", "email_raw", "created_raw",
    ]
    return master[cols]

# ----------------------------
# Reporting period logic
# ----------------------------

def prev_year_month(y: int, m: int) -> Tuple[int, int]:
    return (y - 1, 12) if m == 1 else (y, m - 1)

def last_business_day(y: int, m: int) -> date:
    last_day = calendar.monthrange(y, m)[1]
    d = date(y, m, last_day)
    while d.weekday() >= 5:  # Sat/Sun
        d = date.fromordinal(d.toordinal() - 1)
    return d

def reporting_month_for_run(run_dt: datetime, cutoff_day: int = 15) -> Tuple[int, int]:
    """
    Monthly reporting period selection:
    - If run day <= cutoff_day: previous month
    - Else: current month ONLY if run day is last business day of month; otherwise previous month
    """
    y, m, d = run_dt.year, run_dt.month, run_dt.day

    if d <= cutoff_day:
        return prev_year_month(y, m)

    if run_dt.date() == last_business_day(y, m):
        return (y, m)

    return prev_year_month(y, m)

# ----------------------------
# Filtering helpers
# ----------------------------

def quarter_months(q: str) -> List[int]:
    q = q.upper()
    if q not in FISCAL_QUARTERS:
        raise ValueError(f"Quarter must be one of {list(FISCAL_QUARTERS.keys())}")
    return FISCAL_QUARTERS[q]

def filter_month(master: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    df = master.copy()
    df = df[df["year"].notna() & df["month"].notna()]
    return df[(df["year"].astype(int) == int(year)) & (df["month"].astype(int) == int(month))]

def filter_quarter(master: pd.DataFrame, q: str, year: int) -> pd.DataFrame:
    months = quarter_months(q)
    df = master.copy()
    df = df[df["year"].notna() & df["month"].notna()]
    df = df[(df["year"].astype(int) == int(year)) & (df["month"].astype(int).isin(months))]
    return df

def most_recent_month_from_rbw_carpool(master: pd.DataFrame) -> Tuple[int, int]:
    df = master[master["program"].isin([PROGRAM_RBW, PROGRAM_CARPOOL])].copy()
    df = df.dropna(subset=["created_date"])
    if df.empty:
        raise ValueError("No dated records found in RBW/CARPOOL to determine most recent month.")
    max_dt = df["created_date"].max()
    return int(max_dt.year), int(max_dt.month)

def most_recent_month_in_quarter(master: pd.DataFrame, q: str, year: int) -> Tuple[int, int]:
    df = filter_quarter(master[master["program"].isin([PROGRAM_RBW, PROGRAM_CARPOOL])], q, year)
    df = df.dropna(subset=["created_date"])
    if df.empty:
        raise ValueError(f"No RBW/CARPOOL activity found in {q} {year}.")
    max_dt = df["created_date"].max()
    return int(max_dt.year), int(max_dt.month)

# ----------------------------
# Rewards: lunches
# ----------------------------

def lunches_from_trips(n: int) -> int:
    if n < 5:
        return 0
    if 5 <= n <= 8:
        return 1
    if 9 <= n <= 12:
        return 2
    if 13 <= n <= 15:
        return 3
    if 16 <= n <= 19:
        return 4
    return 5  # 20–31 => 5 (cap)

def calculate_lunch_report(month_df: pd.DataFrame) -> pd.DataFrame:
    df = month_df[month_df["program"].isin([PROGRAM_RBW, PROGRAM_CARPOOL])].copy()
    df = df[df["badge_id"].astype(str).str.len() > 0]

    trip_counts = (
        df.groupby(["badge_id"], as_index=False)
          .agg(trips=("badge_id", "size"),
               name=("name", "first"),
               email=("email", "first"))
    )

    trip_counts["lunches"] = trip_counts["trips"].apply(lunches_from_trips)

    # -------------------------------------------------
    # REMOVE PEOPLE WITH ZERO LUNCHES
    # -------------------------------------------------
    trip_counts = trip_counts[trip_counts["lunches"] > 0]

    # sort after filtering
    trip_counts = (
        trip_counts
        .sort_values(["lunches", "trips"], ascending=[False, False])
        .reset_index(drop=True)
    )

    return trip_counts[["name", "badge_id", "email", "trips", "lunches"]]

# ----------------------------
# Drawings: winners
# ----------------------------

def unique_pool(df: pd.DataFrame) -> pd.DataFrame:
    x = df.copy()
    x["key"] = x["badge_id"]
    x.loc[x["key"].astype(str).str.len() == 0, "key"] = x["email"]
    x = x[x["key"].astype(str).str.len() > 0]
    x = x.drop_duplicates(subset=["key"]).reset_index(drop=True)
    return x

def draw_winners(pool_df: pd.DataFrame, n: int, seed: int) -> pd.DataFrame:
    if len(pool_df) == 0 or n <= 0:
        return pool_df.iloc[0:0].copy()
    n = min(n, len(pool_df))
    return pool_df.sample(n=n, replace=False, random_state=seed).copy()

def run_monthly_drawing(month_df: pd.DataFrame, exclude_keys: set, seed: int) -> Tuple[pd.DataFrame, set]:
    participants = month_df[month_df["program"].isin([PROGRAM_RBW, PROGRAM_CARPOOL])].copy()
    pool = unique_pool(participants)
    pool = pool[~pool["key"].isin(exclude_keys)].reset_index(drop=True)

    winners = draw_winners(pool, n=8, seed=seed)
    winners["program_award"] = "MONTHLY_RBW_CARPOOL"

    new_exclude = set(winners["key"].tolist())
    return winners[["program_award", "name", "badge_id", "email", "key"]], exclude_keys.union(new_exclude)

def run_rad_drawing(quarter_df: pd.DataFrame, exclude_keys: set, seed: int) -> Tuple[pd.DataFrame, set]:
    rad = quarter_df[quarter_df["program"] == PROGRAM_RAD].copy()
    pool = unique_pool(rad)
    pool = pool[~pool["key"].isin(exclude_keys)].reset_index(drop=True)

    winners = draw_winners(pool, n=4, seed=seed)
    winners["program_award"] = "QUARTERLY_RAD"

    new_exclude = set(winners["key"].tolist())
    return winners[["program_award", "name", "badge_id", "email", "key"]], exclude_keys.union(new_exclude)

def run_afv_drawing(master: pd.DataFrame, exclude_keys: set, seed: int) -> Tuple[pd.DataFrame, set]:
    afv = master[master["program"] == PROGRAM_AFV].copy()
    pool = unique_pool(afv)
    pool = pool[~pool["key"].isin(exclude_keys)].reset_index(drop=True)

    winners = draw_winners(pool, n=2, seed=seed)
    winners["program_award"] = "QUARTERLY_AFV"

    new_exclude = set(winners["key"].tolist())
    return winners[["program_award", "name", "badge_id", "email", "key"]], exclude_keys.union(new_exclude)

# ----------------------------
# Seeds / Audit
# ----------------------------

def build_run_seed(run_type: str, year: int, month: Optional[int], quarter: Optional[str]) -> int:
    """
    Deterministic seed stable across machines/sessions (unlike Python's hash()).
    """
    base = f"{run_type}|{year}|{month or ''}|{quarter or ''}"
    digest = hashlib.sha256(base.encode("utf-8")).hexdigest()
    return int(digest[:8], 16)  # 32-bit-ish

# ----------------------------
# Output writing
# ----------------------------


def _last_day_of_month(y: int, m: int) -> date:
    return date(y, m, calendar.monthrange(y, m)[1])


def _add_month(y: int, m: int, delta: int = 1) -> tuple[int, int]:
    # delta=1 means next month
    nm = m + delta
    ny = y
    while nm > 12:
        nm -= 12
        ny += 1
    while nm < 1:
        nm += 12
        ny -= 1
    return ny, nm


def write_lunch_checkoff_xlsx(
    outdir: str,
    lunch_report: pd.DataFrame,
    period_year: int,
    period_month: int,
    site_name: str = "Chandler",
    filename: str = "lunch_checkoff.xlsx",
) -> str:
    """
    Creates a cafeteria-friendly checkoff sheet:
    Name | # Lunches | 1 | 2 | 3 | 4 | 5
    Slots above allowed lunches are blacked out.
    """
    os.makedirs(outdir, exist_ok=True)

    # Expire = last day of NEXT month after the reporting month (matches your screenshot behavior)
    ey, em = _add_month(period_year, period_month, delta=1)
    expire_dt = _last_day_of_month(ey, em)
    expire_str = expire_dt.strftime("%m/%d/%Y")

    # Month label (e.g., "February")
    month_name = calendar.month_name[period_month]

    # Ensure expected columns exist
    if not {"name", "lunches"}.issubset(set(lunch_report.columns)):
        raise ValueError("lunch_report must contain columns: 'name' and 'lunches'")

    df = lunch_report.copy()
    df["lunches"] = pd.to_numeric(df["lunches"], errors="coerce").fillna(0).astype(int)
    df = df.sort_values(["name"], ascending=True).reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} Lunches"

    # --- Styling ---
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    small_bold = Font(bold=True, size=10)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    black_fill = PatternFill("solid", fgColor="000000")
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Top title rows (like your screenshot) ---
    # Row 1: site
    ws["A1"] = site_name
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = left

    # Row 2: month label left, expire right
    ws["A2"] = f"{month_name} TRP lunches"
    ws["A2"].font = small_bold
    ws["A2"].alignment = left
    ws.merge_cells("A2:D2")

    ws["E2"] = f"ALL LUNCHES EXPIRE ON {expire_str}"
    ws["E2"].font = small_bold
    ws["E2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("E2:G2")

    # Row 3 blank spacer
    ws.merge_cells("A3:G3")

    # Row 4: “Lunches” header spanning slot columns
    ws["C4"] = "Lunches"
    ws["C4"].font = bold
    ws["C4"].alignment = center
    ws.merge_cells("C4:G4")

    # Row 5: table headers
    headers = ["Name", "# Lunches", "1", "2", "3", "4", "5"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = bold
        cell.alignment = center if col_idx >= 2 else left
        cell.fill = header_fill
        cell.border = border

    # --- Data rows start at row 6 ---
    start_row = 6
    for i, row in df.iterrows():
        r = start_row + i
        name = str(row["name"]).strip()
        lunches = int(row["lunches"])

        ws.cell(r, 1, name).alignment = left
        ws.cell(r, 2, lunches).alignment = center

        # borders for name + lunches
        ws.cell(r, 1).border = border
        ws.cell(r, 2).border = border

        # Slot columns C..G are 1..5
        for slot in range(1, 6):
            c = 2 + slot  # slot 1 => col 3 (C)
            cell = ws.cell(r, c, "")
            cell.alignment = center
            cell.border = border

            # If slot is beyond their allowed lunches, black it out
            if slot > lunches:
                cell.fill = black_fill

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 6

    # Row heights for nicer spacing
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18

    # Freeze headers so staff can scroll
    ws.freeze_panes = "A6"

    out_path = os.path.join(outdir, filename)
    wb.save(out_path)
    return out_path

def ensure_outputs_dir(outdir: str) -> None:
    os.makedirs(outdir, exist_ok=True)

def write_outputs(outdir: str, cleaned_master: pd.DataFrame, lunch_report: pd.DataFrame, winners_report: pd.DataFrame, run_log: Dict) -> None:
    ensure_outputs_dir(outdir)

    cleaned_master.to_csv(os.path.join(outdir, "cleaned_master.csv"), index=False)
    lunch_report.to_csv(os.path.join(outdir, "lunch_report.csv"), index=False)
    winners_report.to_csv(os.path.join(outdir, "winner_report.csv"), index=False)

    with open(os.path.join(outdir, "run_log.json"), "w", encoding="utf-8") as f:
        json.dump(run_log, f, indent=2, default=str)

def assert_expected_outputs(outdir: str) -> None:
    missing = []
    for fname in EXPECTED_OUTPUT_FILES:
        p = os.path.join(outdir, fname)
        if not os.path.exists(p):
            missing.append(p)
    if missing:
        raise FileNotFoundError("Expected output files missing:\n" + "\n".join(missing))